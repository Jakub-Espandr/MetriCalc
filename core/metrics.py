import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from sklearn.metrics import precision_score, recall_score, f1_score, accuracy_score, cohen_kappa_score
from .translations import TRANSLATIONS, get_class_names

def compute_metrics(df, language='cs'):
    """Compute metrics from confusion matrix data"""
    df = df.copy()
    df.columns = df.columns.astype(str)

    # Find C_* columns (handle cases like "C_1 - nezasazena uroda")
    c_columns = [col for col in df.columns if col.startswith('C_') and '_' in col]
    if not c_columns:
        raise ValueError("No C_* columns found in the CSV file")
    
    # Sort columns to ensure C_1, C_2, C_3, etc. order
    c_columns.sort(key=lambda x: int(x.split('_')[1].split()[0]) if x.split('_')[1].split()[0].isdigit() else 0)
    
    # Filter for rows that have ClassValue matching the C_* pattern
    # Look for ClassValue entries that start with C_ and contain a number
    class_values = []
    for col in c_columns:
        class_num = col.split('_')[1].split()[0]  # Extract number from C_1, C_2, etc.
        if class_num.isdigit():
            class_values.append(f'C_{class_num}')
    
    if not class_values:
        raise ValueError("No valid class values found in ClassValue column")
    
    df_cm = df[df['ClassValue'].astype(str).str.startswith(tuple(class_values))]
    
    if df_cm.empty:
        raise ValueError("No rows found with matching ClassValue entries")

    # Handle decimal commas and convert to numeric
    for col in c_columns:
        df_cm[col] = df_cm[col].astype(str).str.replace(',', '.').astype(float).astype(int)

    # Create confusion matrix from the C_* columns
    cm = df_cm[c_columns].to_numpy()
    
    if cm.size == 0 or cm.shape[0] == 0:
        raise ValueError("Confusion matrix is empty")

    # Create y_true and y_pred arrays
    y_true = []
    y_pred = []
    
    for i, row in enumerate(cm):
        # Add true labels (class i repeated by the sum of that row)
        row_sum = int(row.sum())
        y_true.extend([i] * row_sum)
        
        # Add predicted labels
        for j, count in enumerate(row):
            y_pred.extend([j] * int(count))
    
    if not y_true or not y_pred:
        raise ValueError("No valid predictions found in the data")

    # Calculate metrics
    precision = precision_score(y_true, y_pred, average=None, zero_division=0)
    recall = recall_score(y_true, y_pred, average=None, zero_division=0)
    f1 = f1_score(y_true, y_pred, average=None, zero_division=0)
    accuracy = round(accuracy_score(y_true, y_pred), 3)
    kappa = round(cohen_kappa_score(y_true, y_pred), 3)

    avg_precision = round(precision_score(y_true, y_pred, average='macro', zero_division=0), 3)
    avg_recall = round(recall_score(y_true, y_pred, average='macro', zero_division=0), 3)
    avg_f1 = round(f1_score(y_true, y_pred, average='macro', zero_division=0), 3)

    # Generate class names based on the number of classes found
    class_names = get_class_names(len(c_columns), language)
    
    # Create results for each class
    results = []
    for i in range(len(c_columns)):
        if i < len(precision):
            results.append([
                class_names[i] if i < len(class_names) else f"Class {i+1}", 
                round(precision[i], 3), 
                round(recall[i], 3), 
                round(f1[i], 3), 
                accuracy, 
                kappa
            ])
    
    # Add average row
    results.append([
        class_names[-1] if len(class_names) > len(c_columns) else "Average", 
        avg_precision, 
        avg_recall, 
        avg_f1, 
        accuracy, 
        kappa
    ])
    
    return results

def export_to_excel(input_path, output_path, language='cs'):
    """Export metrics to Excel file"""
    try:
        df = pd.read_csv(input_path, sep=';')
        metrics = compute_metrics(df, language)
        wb = Workbook()
        sheetname = Path(input_path).stem

        # Metrics sheet
        ws1 = wb.active
        ws1.title = f"{TRANSLATIONS[language]['excel_metrics_sheet']}_{sheetname}"
        headers = TRANSLATIONS[language]['headers']
        for col, header in enumerate(headers, start=1):
            ws1.cell(row=1, column=col).value = header
        for row_i, row_data in enumerate(metrics, start=2):
            for col_i, value in enumerate(row_data, start=1):
                ws1.cell(row=row_i, column=col_i).value = value

        # Data sheet
        ws2 = wb.create_sheet(f"{TRANSLATIONS[language]['excel_data_sheet']}_{sheetname}")
        ws2.append(list(df.columns))
        for r in df.itertuples(index=False):
            ws2.append(list(r))

        wb.save(output_path)
        return True, None, (metrics, df)
    except Exception as e:
        return False, str(e), None

def add_to_workbook(wb, input_path, metrics, df, language='cs'):
    """Add data to existing workbook"""
    try:
        sheetname = Path(input_path).stem
        
        # Metrics sheet
        ws1 = wb.create_sheet(f"{TRANSLATIONS[language]['excel_metrics_sheet']}_{sheetname}")
        headers = TRANSLATIONS[language]['headers']
        for col, header in enumerate(headers, start=1):
            ws1.cell(row=1, column=col).value = header
        for row_i, row_data in enumerate(metrics, start=2):
            for col_i, value in enumerate(row_data, start=1):
                ws1.cell(row=row_i, column=col_i).value = value

        # Data sheet
        ws2 = wb.create_sheet(f"{TRANSLATIONS[language]['excel_data_sheet']}_{sheetname}")
        ws2.append(list(df.columns))
        for r in df.itertuples(index=False):
            ws2.append(list(r))
        
        return True, None
    except Exception as e:
        return False, str(e) 